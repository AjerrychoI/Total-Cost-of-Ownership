import pandas as pd
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter

def create_tco_professional_usd(file_name):  
    # 輸入參數 (美元單位)  
    input_params = {  
        "Parameter": [  
            "Initial Cost (初始成本)",   
            "Lifetime (使用年限)",   
            "Annual Maintenance (每年維護成本)",   
            "Annual Operating (每年營運成本)",   
            "Salvage Value (殘值)",   
            "Inflation Rate (通膨率)"  
        ],  
        "Value": [10000, 5, 2000, 3000, 1000, 0.02],  
        "Unit": ["USD", "Years", "USD/Year", "USD/Year", "USD", "Percentage"]  
    }  

    # 參數提取  
    lifetime = input_params["Value"][1]  
    initial_cost = input_params["Value"][0]  
    maintenance = input_params["Value"][2]  
    operating = input_params["Value"][3]  
    salvage = input_params["Value"][4]  
    inflation = input_params["Value"][5]  

    # 成本計算  
    yearly_data = []  
    cumulative = 0  
    maintenance_adj = maintenance  
    operating_adj = operating  

    for year in range(1, lifetime + 1):  
        capex = initial_cost if year == 1 else 0  
        
        if year > 1:  
            maintenance_adj *= (1 + inflation)  
            operating_adj *= (1 + inflation)  
        
        opex = maintenance_adj + operating_adj  
        annual_total = capex + opex  
        cumulative += annual_total  

        yearly_data.append({  
            "Year (年份)": year,  
            "CapEx ($) (資本支出)": capex,  
            "Maintenance ($) (維護成本)": round(maintenance_adj, 2),  
            "Operating ($) (營運成本)": round(operating_adj, 2),  
            "OpEx Total ($) (營運總成本)": round(opex, 2),  
            "Annual Total ($) (年度總成本)": round(annual_total, 2),  
            "Cumulative ($) (累積成本)": round(cumulative, 2)  
        })  

    # 財務總結  
    summary_data = {  
        "Metric (指標)": [  
            "Total CapEx (總資本支出)",   
            "Total OpEx (總營運成本)",   
            "Salvage Deduction (殘值扣除)",   
            "Net TCO (淨總擁有成本)"  
        ],  
        "Amount ($) (金額)": [  
            initial_cost,  
            round(sum(row["OpEx Total ($) (營運總成本)"] for row in yearly_data), 2),  
            salvage,  
            round(cumulative - salvage, 2)  
        ]  
    }  

    # 生成 DataFrames  
    df_input = pd.DataFrame(input_params)  
    df_yearly = pd.DataFrame(yearly_data)  
    df_summary = pd.DataFrame(summary_data)  

    # 輸出 Excel (美元格式)  
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:  
        df_input.to_excel(writer, sheet_name="Input Parameters (輸入參數)", index=False)  
        df_yearly.to_excel(writer, sheet_name="Cost Breakdown (成本明細)", index=False)  
        df_summary.to_excel(writer, sheet_name="Financial Summary (財務總結)", index=False)  

        # 設定格式  
        workbook = writer.book  
        for sheet_name in writer.sheets:  
            sheet = writer.sheets[sheet_name]  
            for col_idx, col in enumerate(sheet.iter_cols(), start=1):  
                col_letter = get_column_letter(col_idx)  
                for cell in col:  
                    if isinstance(cell.value, (int, float)):  
                        # 設定千分位符號格式
                        cell.number_format = '#,##0.00' if isinstance(cell.value, float) else '#,##0'
            sheet.freeze_panes = "A2"  # 凍結標題列  

    print(f"Professional TCO Report Generated: {file_name}")  

# 執行範例  
create_tco_professional_usd("TCO_Analysis_Template.xlsx")